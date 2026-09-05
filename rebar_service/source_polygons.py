from __future__ import annotations

import re
import warnings
from io import BytesIO
from numbers import Real
from typing import Any, Callable, Iterable, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from A101.read_dxf import extract_polygons_from_bytes


class SourcePolygonsError(ValueError):
    pass


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").replace("\n", " ").split()).strip().lower()


def _is_axis_header(value: str, axis: str) -> bool:
    value = value.strip().lower()
    return value == axis or value.startswith(f"{axis} ") or value.startswith(f"{axis}(")


def _header_columns(
    sheet: Worksheet,
    *,
    label: str,
    required: Mapping[str, Callable[[str], bool]],
) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        normalized = [_normalize_header(value) for value in row]
        found: dict[str, int] = {}
        for key, predicate in required.items():
            for index, value in enumerate(normalized):
                if predicate(value):
                    found[key] = index
                    break
        if len(found) == len(required):
            return row_number, found
        if row_number >= 100:
            break
    names = ", ".join(required)
    raise SourcePolygonsError(f"{label}: required header row not found ({names})")


def _as_int_id(value: Any, *, label: str) -> int:
    if isinstance(value, bool) or value is None:
        raise SourcePolygonsError(f"{label} must be an integer")
    if isinstance(value, Real):
        number = float(value)
        if not number.is_integer():
            raise SourcePolygonsError(f"{label} must be an integer")
        return int(number)
    text = str(value).strip()
    try:
        number = float(text.replace(",", "."))
    except ValueError as exc:
        raise SourcePolygonsError(f"{label} must be an integer") from exc
    if not number.is_integer():
        raise SourcePolygonsError(f"{label} must be an integer")
    return int(number)


def _as_float(value: Any, *, label: str) -> float:
    if isinstance(value, bool) or value is None:
        raise SourcePolygonsError(f"{label} must be numeric")
    if isinstance(value, Real):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        raise SourcePolygonsError(f"{label} must be numeric")
    try:
        return float(text)
    except ValueError as exc:
        raise SourcePolygonsError(f"{label} must be numeric") from exc


def _row_value(row: Sequence[Any], index: int) -> Any:
    return row[index] if index < len(row) else None


def _open_sheet(content: bytes, *, label: str) -> tuple[Any, Worksheet]:
    if not content:
        raise SourcePolygonsError(f"{label}: XLSX file is empty")
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Cannot parse header or footer.*",
                category=UserWarning,
            )
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SourcePolygonsError(f"{label}: failed to read XLSX: {exc}") from exc
    if not workbook.worksheets:
        workbook.close()
        raise SourcePolygonsError(f"{label}: workbook has no worksheets")
    return workbook, workbook.worksheets[0]


def _parse_nodes(content: bytes) -> dict[int, tuple[float, float, float]]:
    workbook, sheet = _open_sheet(content, label="nodes")
    try:
        header_row, columns = _header_columns(
            sheet,
            label="nodes",
            required={
                "node": lambda value: value.replace(" ", "") in {"№узла", "номерузла"},
                "x": lambda value: _is_axis_header(value, "x"),
                "y": lambda value: _is_axis_header(value, "y"),
                "z": lambda value: _is_axis_header(value, "z"),
            },
        )
        nodes: dict[int, tuple[float, float, float]] = {}
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            raw_id = _row_value(row, columns["node"])
            if raw_id is None or str(raw_id).strip() == "":
                continue
            node_id = _as_int_id(raw_id, label=f"nodes row {row_number}: node id")
            if node_id in nodes:
                raise SourcePolygonsError(f"Duplicate node id {node_id}")
            x = _as_float(_row_value(row, columns["x"]), label=f"node {node_id}: X")
            y = _as_float(_row_value(row, columns["y"]), label=f"node {node_id}: Y")
            z = _as_float(_row_value(row, columns["z"]), label=f"node {node_id}: Z")
            nodes[node_id] = (round(x * 1000.0, 9), round(y * 1000.0, 9), round(z * 1000.0, 9))
        if not nodes:
            raise SourcePolygonsError("nodes: no node rows found")
        return nodes
    finally:
        workbook.close()


def _parse_node_list(value: Any, *, element_id: int) -> list[int]:
    if value is None:
        raise SourcePolygonsError(f"element {element_id}: node list is missing")
    if isinstance(value, (list, tuple)):
        raw_values: Iterable[Any] = value
    else:
        text = str(value).strip()
        raw_values = [part for part in re.split(r"[,;\s]+", text) if part]
    node_ids = [_as_int_id(item, label=f"element {element_id}: node id") for item in raw_values]
    if len(node_ids) < 3 or len(set(node_ids)) < 3:
        raise SourcePolygonsError(f"element {element_id}: polygon must reference at least 3 distinct nodes")
    return node_ids


def _parse_elements(content: bytes, nodes: Mapping[int, tuple[float, float, float]]) -> list[tuple[int, list[int]]]:
    workbook, sheet = _open_sheet(content, label="elements")
    try:
        header_row, columns = _header_columns(
            sheet,
            label="elements",
            required={
                "element": lambda value: value.startswith("№ элем") or value.replace(" ", "") in {"№элем", "№элемента"},
                "nodes": lambda value: "№№ узлов" in value or value.replace(" ", "") in {"№№узлов", "узлы"},
            },
        )
        elements: list[tuple[int, list[int]]] = []
        seen: set[int] = set()
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            raw_id = _row_value(row, columns["element"])
            if raw_id is None or str(raw_id).strip() == "":
                continue
            element_id = _as_int_id(raw_id, label=f"elements row {row_number}: element id")
            if element_id in seen:
                raise SourcePolygonsError(f"Duplicate element id {element_id}")
            node_ids = _parse_node_list(_row_value(row, columns["nodes"]), element_id=element_id)
            for node_id in node_ids:
                if node_id not in nodes:
                    raise SourcePolygonsError(f"element {element_id}: unknown node {node_id}")
            seen.add(element_id)
            elements.append((element_id, node_ids))
        if not elements:
            raise SourcePolygonsError("elements: no element rows found")
        return elements
    finally:
        workbook.close()


def _parse_loads(content: bytes, *, load_column: int) -> dict[int, float]:
    if int(load_column) not in {1, 2, 3, 4}:
        raise SourcePolygonsError("load_column must be in range 1..4")
    selected_name = f"AS{int(load_column)}"
    workbook, sheet = _open_sheet(content, label="loads")
    try:
        header_row, columns = _header_columns(
            sheet,
            label="loads",
            required={
                "element": lambda value: value == "элемент",
                "as1": lambda value: value == "as1",
                "as2": lambda value: value == "as2",
                "as3": lambda value: value == "as3",
                "as4": lambda value: value == "as4",
            },
        )
        selected_index = columns[f"as{int(load_column)}"]
        loads: dict[int, float] = {}
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            raw_id = _row_value(row, columns["element"])
            if raw_id is None or str(raw_id).strip() == "":
                continue
            element_id = _as_int_id(raw_id, label=f"loads row {row_number}: element id")
            if element_id in loads:
                # The LIRA export contains two rows per element. The first one is authoritative.
                continue
            value = _row_value(row, selected_index)
            loads[element_id] = _as_float(value, label=f"{selected_name} for element {element_id}")
        if not loads:
            raise SourcePolygonsError("loads: no load rows found")
        return loads
    finally:
        workbook.close()


def source_polygons_from_xlsx(
    nodes_content: bytes,
    elements_content: bytes,
    loads_content: bytes,
    load_column: int,
) -> dict[str, Any]:
    """Convert the three LIRA XLSX exports into the canonical polygon task input."""
    if int(load_column) not in {1, 2, 3, 4}:
        raise SourcePolygonsError("load_column must be in range 1..4")

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer.*",
            category=UserWarning,
        )
        nodes = _parse_nodes(nodes_content)
        elements = _parse_elements(elements_content, nodes)
        loads = _parse_loads(loads_content, load_column=int(load_column))

        polygons: list[dict[str, Any]] = []
        for element_id, node_ids in elements:
            if element_id not in loads:
                raise SourcePolygonsError(f"Missing load for element {element_id}")
            polygons.append(
                {
                    "points": [[float(nodes[node_id][0]), float(nodes[node_id][1])] for node_id in node_ids],
                    "load": float(loads[element_id]),
                }
            )
        return {"kind": "polygons", "units": "mm", "polygons": polygons}


def source_polygons_from_input(source_input: Mapping[str, Any]) -> list[dict[str, Any]]:
    """Return task source polygons in millimetres, independently of the original source format."""
    if not isinstance(source_input, Mapping):
        raise SourcePolygonsError("Source input is invalid")

    kind = source_input.get("kind")
    if kind == "polygons":
        scale = 1000.0 if source_input.get("units", "mm") == "m" else 1.0
        result: list[dict[str, Any]] = []
        for index, item in enumerate(source_input.get("polygons", []) or []):
            if not isinstance(item, Mapping):
                raise SourcePolygonsError(f"Invalid polygon #{index}")
            try:
                points = [
                    [float(point[0]) * scale, float(point[1]) * scale]
                    for point in item["points"]
                ]
                load = float(item["load"])
            except (KeyError, TypeError, ValueError, IndexError) as exc:
                raise SourcePolygonsError(f"Invalid polygon #{index}") from exc
            if len(points) < 3:
                raise SourcePolygonsError(f"Invalid polygon #{index}")
            row: dict[str, Any] = {"points": points, "load": load}
            if item.get("color") is not None:
                row["color"] = int(item["color"])
            result.append(row)
        if not result:
            raise SourcePolygonsError("Source polygons are empty")
        return result

    if kind != "dxf":
        raise SourcePolygonsError("Source polygons are available only for DXF or polygon tasks")

    content = source_input.get("content")
    if not isinstance(content, (bytes, bytearray, memoryview)):
        raise SourcePolygonsError("DXF source content is missing")

    try:
        mosaic = extract_polygons_from_bytes(bytes(content))
    except Exception as exc:
        raise SourcePolygonsError(f"Failed to parse source DXF: {exc}") from exc

    return [
        {
            "points": [[float(x), float(y)] for x, y in poly["points"]],
            "color": int(poly["color"]),
            "load": float(poly["load"]),
        }
        for poly in mosaic
    ]
